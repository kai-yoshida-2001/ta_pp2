#!/usr/bin/env python3

import openpyxl, os
from pathlib import Path

# Q1.'Workbook'モジュールで新規ファイルを作成しよう
wb = openpyxl.Workbook()
output_path = Path('./output/normal_example.xlsx')

sheet1 = wb.active
# Q2.'title'メソッドを使い，シート名を定義しよう
sheet1.title = 'origin'

sheet1.append(['名前', '田中', '佐藤', '高橋'])
sheet1.append(['年齢', 20, 21, 22])
sheet1.append(['趣味', '読書', '映画鑑賞', '料理'])

print('===元データ===')
for row in sheet1.iter_rows(values_only=True):
    print(row)

# Q3.'create_sheet'メソッドを使い，2つ目のシートを'swap_rows_and_cols'という名前で作成しよう
'''
ヒント1：
新規作成した'Workbook'を定義している変数を見つけよう
ヒント2：
シート名は'title'で定義することができます．
'''
sheet2 = wb.create_sheet(title='swap_rows_and_cols')

data = list(sheet1.iter_rows(values_only=True))
transposed_data = zip(*data)

for row in transposed_data:
    sheet2.append(row)

print('===転置データ===')
for row in sheet2.iter_rows(values_only=True):
    print(row)
    
os.makedirs(output_path.parent, exist_ok=True)
# Q4.'save'メソッドを使い，'output'ディレクトリに保存しよう
'''
ヒント1：
保存する'workbook'を定義した変数名を思い出そう
ヒント2：
'output'ディレクトリと保存する際のファイル名を
定義している変数を見つけよう
'''
wb.save(output_path)
# Q5.4問目と同じパスを，{}内に記述しよう
print(f'ワークシートを{output_path}に保存しました')
