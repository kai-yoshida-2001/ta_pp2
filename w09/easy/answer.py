#!/usr/bin/env python3

import openpyxl, os
from pathlib import Path

input_path = Path('./input/easy_example.xlsx')
# Q1.'load_workbook'メソッドを呼び出して，Excelファイルを読み込もう
wb = openpyxl.load_workbook(input_path)

print('======')
# Q2.'sheetnames'メソッドを呼び出して，'workbook'のシート情報を取得しよう
'''
ヒント1：
'workbook'という変数は定義されていないので，
'workbook'を引数に指定するとエラーを検出してしまいます．
ヒント2：
'workbook'という情報を定義している変数を見つけ，
メソッドの前に定義しましょう．
'''
print(f'取得したシート名一覧：{wb.sheetnames}')

# Q3.'workbook'のシート名'Sheet1'を指定しよう（'workbook'に注意しよう）
select_sheet = wb['Sheet1']
print('======')
# Q4.'title'メソッドを呼び出して，指定したシートのタイトルを取得しよう
'''
ヒント：
指定したシート名を定義している変数を見つけよう
'''
print(f'指定されたシート：{select_sheet.title}')

read_num_cell = 'A1'
# Q5.'value'メソッドを使い，指定したシート内のセル情報を取得しよう
'''
ヒント1：
指定したシート名を定義している変数を思い出そう
ヒント2：
情報を取得したいセル番号を定義した変数を見つけよう
'''
cell_data = select_sheet[read_num_cell].value
print('======')
print(f'{read_num_cell}セルの値：{cell_data}')

write_num_cell = 'A4'
# Q6.指定したシート内のセルにテキストを書き込もう
'''
ヒント1：
指定したシート名を定義している変数を思い出そう
ヒント2：
書き込みたいセル番号を定義した変数を見つけよう
'''
select_sheet[write_num_cell] = '業種'

output_path = Path('./output/output_example.xlsx')
os.makedirs(output_path.parent, exist_ok=True)

# Q7.'save'メソッドを使い，'output'ディレクトリにファイルを保存しよう
'''
ヒント：
'output'ディレクトリや保存の際のファイル名をパスで定義している
変数を見つけよう
'''
wb.save(output_path)
print('======')
print(f'ファイルを保存しました：{output_path}')
